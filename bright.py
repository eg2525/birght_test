import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill

# 科学記数法の表示を無効にする
pd.set_option('display.float_format', '{:.2f}'.format)

def process_excel(file_sta, file_now, lower_limit, percentage_diff):
    ex_sta = pd.ExcelFile(file_sta)
    df_sta = pd.read_excel(ex_sta, header=1, sheet_name=None)

    # 処理結果を保存する辞書を用意
    results = {}

    # 各シートに対して処理を実行
    for key in df_sta.keys():
        # 決算書表示名と期間累計の位置を取得
        start_col = df_sta[key].columns.get_loc('決算書表示名') + 1
        end_col = df_sta[key].columns.get_loc('期間累計') - 1  # 期間累計の2つ前の列を指す

        # 決算書表示名の次の列から期間累計の2つ前の列までを選択して平均を計算
        result = df_sta[key].iloc[:, start_col:end_col].mean(axis=1)

        # 四捨五入して整数に変換
        result = result.round(0).astype(int)

        # インデックスを勘定科目に設定
        result.index = df_sta[key]['勘定科目']

        # 結果を辞書に保存
        results[key] = result

    ex_now = pd.ExcelFile(file_now)
    df_now = pd.read_excel(ex_now, header=1, sheet_name=None)

    # resultsとdf_nowの照合と値の追加
    for key in results.keys():
        if key in df_now:
            # 前年平均の列を追加
            df_now[key]['前年平均'] = None
            
            for idx, row in df_now[key].iterrows():
                account = row['勘定科目']
                if account in results[key].index:
                    df_now[key].at[idx, '前年平均'] = results[key][account]

    # 必要であれば、結果を新しいExcelファイルに保存
    with pd.ExcelWriter('月次推移_損益計算書_更新.xlsx') as writer:
        for key in df_now:
            df_now[key].to_excel(writer, sheet_name=key, index=False)

    # 保存したExcelファイルを読み込む
    file_path = '月次推移_損益計算書_更新.xlsx'
    workbook = load_workbook(file_path)

    # 装飾を設定する行のリスト
    decorated_rows = ["売上高", "売上原価", "売上総損益", "販売費及び一般管理費",
                      "営業損益", "営業外利益", "営業外損失", "経常損益", "特別利益",
                      "特別損失", "税引前当期純利益"]

    # 太字と罫線のスタイル設定
    bold_font = Font(bold=True)
    thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    color_fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    # 全シートに対して処理を実行
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # 列幅の調整
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if sheet.column_dimensions[column].width < adjusted_width:
                sheet.column_dimensions[column].width = adjusted_width
        
        # ヘッダーの装飾を削除
        for cell in sheet[1]:  # assuming header is in the second row
            cell.font = Font(bold=False)
            cell.border = Border()

        # 1行目にタイトルを挿入
        title = f"月次推移：損益計算書(表示単位：円) - {sheet_name}"
        sheet.insert_rows(1)
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True)

        # 装飾を設定する行を太文字にし、上下に罫線を引く
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value in decorated_rows:
                    for cell in sheet[cell.row]:
                        cell.font = bold_font
                        cell.border = thin_border

        # 数値が含まれるセルに桁区切りの数値書式を設定
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        
        # '全体'シートの場合、色付けを追加
        if sheet.title == '全体':
            # '前年平均'列と'決算書表示名'から'期間累計'の間の数値を比較
            average_col = None
            for col in sheet.iter_cols(min_row=2, max_row=2):
                for cell in col:
                    if cell.value == '前年平均':
                        average_col = cell.column
                        break
                if average_col is not None:
                    break

            if average_col is not None:
                # '決算書表示名'と'期間累計'の位置を取得
                start_col = None
                end_col = None
                for col in sheet.iter_cols(min_row=2, max_row=2):
                    for cell in col:
                        if cell.value == '決算書表示名':
                            start_col = cell.column
                        elif cell.value == '期間累計':
                            end_col = cell.column
                        if start_col and end_col:
                            break
                    if start_col and end_col:
                        break

                if start_col and end_col:
                    skip_rest = False  # フラグを初期化
                    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
                        account_name = row[0].value  # '勘定科目'列の値を取得
                        if account_name == '営業損益':
                            skip_rest = True  # '営業損益'が見つかったらフラグを設定
                        if skip_rest:
                            break  # フラグが設定されたらループを抜ける

                        average_value = row[average_col - 1].value
                        if average_value is not None:
                            lower_bound = average_value * (1 - percentage_diff / 100)
                            upper_bound = average_value * (1 + percentage_diff / 100)
                            for cell in row[start_col-1:end_col-1]:
                                if isinstance(cell.value, (int, float)):
                                    difference = abs(cell.value - average_value)
                                    # ±%範囲外かつ差額が下限値以上の場合、セルに色を付ける
                                    if (cell.value < lower_bound or cell.value > upper_bound) and difference >= lower_limit:
                                        cell.fill = color_fill

    # 編集したExcelファイルを保存
    output_file_path = f'月次推移_損益計算書_前期比較(±{lower_limit // 10000}万円).xlsx'
    workbook.save(output_file_path)

    return output_file_path

# Streamlit UI
st.title('freee月次推移表_前年比較')

uploaded_file_sta = st.file_uploader("前年の月次推移損益計算書をアップロードしてください", type="xlsx")
uploaded_file_now = st.file_uploader("現在の月次推移損益計算書をアップロードしてください", type="xlsx")

lower_limit = st.selectbox('下限値を選択してください（円）', [x * 50000 for x in range(1, 21)], index=5)
percentage_diff = st.selectbox('差異の%を選択してください', [x * 5 for x in range(1, 21)], index=4)

if uploaded_file_sta and uploaded_file_now:
    with st.spinner('処理中...'):
        output_file = process_excel(uploaded_file_sta, uploaded_file_now, lower_limit, percentage_diff)
    st.success('処理が完了しました。')
    st.download_button(
        label="処理済みファイルをダウンロード",
        data=open(output_file, "rb").read(),
        file_name=output_file,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
